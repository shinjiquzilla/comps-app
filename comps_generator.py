"""
Comps Table Generator
=====================
Generates a formatted Comparable Company Analysis Excel from a JSON config file.

Usage:
    python comps_generator.py <config.json> [output.xlsx]

JSON config format:
{
    "title": "Comparable Company Analysis (Comps)",
    "date": "2026/2/18",
    "currency": "JPY",
    "unit": "millions",
    "companies": [
        {
            "code": "6763",
            "name": "Teikoku Tsushin Kogyo",
            "sector": "Variable Resistors",
            "accounting": "J-GAAP",
            "fy_end": "Mar",
            "stock_price": 2756,
            "shares_outstanding": 9237,
            "market_cap": 25458,
            "bs_date": "2025/9/30",
            "cash": 12006,
            "total_debt": 77,
            "equity_parent": 27494,
            "equity_ratio": 0.829,
            "rev_ltm": 17316,
            "op_ltm": 1859,
            "ni_ltm": 1757,
            "da_ltm": 771,
            "ebitda_ltm": 2630,
            "rev_forecast": 16800,
            "op_forecast": 1300,
            "ni_forecast": 1200,
            "ebitda_forecast": 2058,
            "dps": 100
        }
    ],
    "notes": [
        "Custom note 1",
        "Custom note 2"
    ]
}
"""

import sys
import io
import json
import statistics

# CLIモードのみstdout/stderrをUTF-8に設定（Streamlit環境では不要・有害）
if __name__ == '__main__':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# Styling constants
# ============================================================
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center')
RIGHT = Alignment(horizontal='right')
WRAP_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def wc(ws, row, col, value, font=None, fill=None, align=None, nf=None):
    """Write a cell with optional formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if nf:
        cell.number_format = nf
    return cell


def safe_div(a, b):
    """Safe division returning None on error."""
    if b is None or b == 0 or a is None:
        return None
    return a / b


def generate_comps(config, output_path):
    """Generate Comps Excel from config dict."""
    companies = config['companies']
    title = config.get('title', 'Comparable Company Analysis (Comps)')
    date_str = config.get('date', '')
    currency = config.get('currency', 'JPY')
    unit = config.get('unit', 'millions')
    custom_notes = config.get('notes', [])

    # Pre-calculate derived values
    n = len(companies)
    evs = []
    ev_ebitda_ltm = []
    ev_ebitda_fwd = []
    per_fwd = []
    pbr_vals = []
    div_yields = []
    op_margins = []
    ebitda_margins = []

    for c in companies:
        mcap = c.get('market_cap', 0) or 0
        debt = c.get('total_debt', 0) or 0
        cash = c.get('cash', 0) or 0
        ev = mcap + debt - cash
        evs.append(ev)

        ebitda_ltm = c.get('ebitda_ltm')
        ebitda_e = c.get('ebitda_forecast')
        ni_e = c.get('ni_forecast')
        equity = c.get('equity_parent')
        price = c.get('stock_price')
        dps = c.get('dps')
        rev = c.get('rev_ltm')
        op = c.get('op_ltm')

        ev_ebitda_ltm.append(safe_div(ev, ebitda_ltm) if ebitda_ltm and ebitda_ltm > 0 else None)
        ev_ebitda_fwd.append(safe_div(ev, ebitda_e) if ebitda_e and ebitda_e > 0 else None)
        per_fwd.append(safe_div(mcap, ni_e) if ni_e and ni_e > 0 else None)
        pbr_vals.append(safe_div(mcap, equity) if equity and equity > 0 else None)
        div_yields.append(safe_div(dps, price) if dps and price and price > 0 else None)

        opm = safe_div(op, rev)
        em = safe_div(c.get('ebitda_ltm'), rev)
        op_margins.append(opm if opm is not None and opm >= 0 else None)
        ebitda_margins.append(em if em is not None and em >= 0 else None)

    # ============================================================
    # Build workbook
    # ============================================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comps"
    ws.column_dimensions['A'].width = 26

    # Define column sections
    sections = []

    info_cols = [
        ("Ticker", lambda i: companies[i].get('code', ''), '@'),
        ("Sector", lambda i: companies[i].get('sector', ''), '@'),
        ("Accounting", lambda i: companies[i].get('accounting', ''), '@'),
        ("FY End", lambda i: companies[i].get('fy_end', ''), '@'),
        ("Stock Price\n({})".format(currency), lambda i: companies[i].get('stock_price'), '#,##0'),
        ("Shares Out.\n('000)", lambda i: companies[i].get('shares_outstanding'), '#,##0'),
        ("Market Cap", lambda i: companies[i].get('market_cap'), '#,##0'),
    ]
    sections.append(("Company Information", info_cols))

    bs_cols = [
        ("BS Date", lambda i: companies[i].get('bs_date', ''), '@'),
        ("Cash &\nDeposits", lambda i: companies[i].get('cash'), '#,##0'),
        ("Total Debt\n(incl. Lease)", lambda i: companies[i].get('total_debt'), '#,##0'),
        ("EV", lambda i: evs[i], '#,##0'),
        ("Equity\n(Parent)", lambda i: companies[i].get('equity_parent'), '#,##0'),
        ("Equity\nRatio", lambda i: companies[i].get('equity_ratio'), '0.0%'),
    ]
    sections.append(("Balance Sheet (Latest)", bs_cols))

    pl_cols = [
        ("Revenue\nLTM", lambda i: companies[i].get('rev_ltm'), '#,##0'),
        ("OP\nLTM", lambda i: companies[i].get('op_ltm'), '#,##0'),
        ("Net Income\nLTM", lambda i: companies[i].get('ni_ltm'), '#,##0'),
        ("D&A\nLTM", lambda i: companies[i].get('da_ltm'), '#,##0'),
        ("EBITDA\nLTM", lambda i: companies[i].get('ebitda_ltm'), '#,##0'),
        ("OP\nMargin", lambda i: op_margins[i], '0.0%'),
        ("EBITDA\nMargin", lambda i: ebitda_margins[i], '0.0%'),
    ]
    sections.append(("P&L - LTM", pl_cols))

    fc_cols = [
        ("Revenue\nFY E", lambda i: companies[i].get('rev_forecast'), '#,##0'),
        ("OP\nFY E", lambda i: companies[i].get('op_forecast'), '#,##0'),
        ("NI\nFY E", lambda i: companies[i].get('ni_forecast'), '#,##0'),
        ("EBITDA\nFY E", lambda i: companies[i].get('ebitda_forecast'), '#,##0'),
    ]
    sections.append(("FY Forecast (Guidance)", fc_cols))

    val_cols = [
        ("EV/EBITDA\n(LTM)", lambda i: ev_ebitda_ltm[i], '0.00x'),
        ("EV/EBITDA\n(FY E)", lambda i: ev_ebitda_fwd[i], '0.00x'),
        ("PER\n(FY E)", lambda i: per_fwd[i], '0.00x'),
        ("PBR", lambda i: pbr_vals[i], '0.00x'),
        ("Div.\nYield", lambda i: div_yields[i], '0.0%'),
    ]
    sections.append(("Valuation Multiples", val_cols))

    # Flatten columns
    all_cols = []
    section_start = {}
    col_idx = 2
    for sec_name, cols in sections:
        section_start[sec_name] = col_idx
        for header, fn, fmt in cols:
            all_cols.append((header, fn, fmt, sec_name))
            col_idx += 1
    total_cols = len(all_cols)

    for j in range(total_cols):
        ws.column_dimensions[get_column_letter(j + 2)].width = 13

    # --- Title ---
    r = 1
    ws.cell(row=r, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=min(total_cols + 1, 15))
    r += 1
    if date_str:
        ws.cell(row=r, column=1, value="Date: {}".format(date_str))
    ws.cell(row=r, column=8, value="Unit: {} {}".format(currency, unit))
    r += 2

    # --- Section headers (merged) ---
    for sec_name, start_col in section_start.items():
        count = len([c for c in all_cols if c[3] == sec_name])
        end_col = start_col + count - 1
        for cc in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=cc)
            cell.fill = SECTION_FILL
            cell.font = SECTION_FONT
            cell.border = THIN_BORDER
        ws.cell(row=r, column=start_col, value=sec_name)
        if count > 1:
            ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=end_col)
    wc(ws, r, 1, "", fill=SECTION_FILL)
    r += 1

    # --- Column headers ---
    wc(ws, r, 1, "Company", font=Font(bold=True, size=10), fill=HEADER_FILL, align=WRAP_CENTER)
    for j, (header, fn, fmt, sec) in enumerate(all_cols):
        wc(ws, r, j + 2, header, font=Font(bold=True, size=9), fill=HEADER_FILL, align=WRAP_CENTER)
    ws.row_dimensions[r].height = 35
    r += 1

    # --- Company data rows ---
    for i, c in enumerate(companies):
        name_label = c.get('name', '')

        wc(ws, r, 1, name_label, font=Font(bold=True, size=10),
           align=Alignment(horizontal='left', vertical='center'))

        for j, (header, fn, fmt, sec) in enumerate(all_cols):
            val = fn(i)
            if val is None:
                wc(ws, r, j + 2, "NM", align=CENTER)
            else:
                wc(ws, r, j + 2, val, nf=fmt, align=RIGHT)
        r += 1

    r += 1

    # --- Peer Statistics ---
    wc(ws, r, 1, "", fill=SECTION_FILL)
    for j in range(total_cols):
        c2 = ws.cell(row=r, column=j + 2)
        c2.fill = SECTION_FILL
        c2.border = THIN_BORDER
    ws.cell(row=r, column=1, value="Summary Statistics").font = SECTION_FONT
    r += 1

    stat_labels = [("Median", statistics.median), ("Mean", statistics.mean),
                   ("Low", min), ("High", max)]

    for stat_name, stat_fn in stat_labels:
        wc(ws, r, 1, stat_name, font=Font(bold=True))
        for j, (header, fn, fmt, sec) in enumerate(all_cols):
            peer_vals = [fn(i) for i in range(n)]
            nums = [v for v in peer_vals if v is not None and isinstance(v, (int, float))]
            is_margin = header in ("OP\nMargin", "EBITDA\nMargin")
            is_multiple = sec == "Valuation Multiples"
            if nums and fmt not in ['@'] and (is_margin or is_multiple):
                wc(ws, r, j + 2, stat_fn(nums), nf=fmt, align=RIGHT)
            else:
                wc(ws, r, j + 2, "", align=CENTER)
        r += 1

    r += 1

    # --- Notes ---
    if custom_notes:
        ws.cell(row=r, column=1, value="Notes:").font = Font(bold=True)
        r += 1
        for note in custom_notes:
            ws.cell(row=r, column=1, value=note)
            r += 1

    # --- Save ---
    wb.save(output_path)
    print("Saved: {}".format(output_path))
    print("Companies: {}, Columns: {}".format(n, total_cols))
    return output_path


# ============================================================
# CLI entry point
# ============================================================
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python comps_generator.py <config.json> [output.xlsx]")
        print("       See docstring for JSON format.")
        sys.exit(1)

    config_path = sys.argv[1]
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    if len(sys.argv) >= 3:
        output = sys.argv[2]
    else:
        output = config_path.replace('.json', '_comps.xlsx')

    generate_comps(config, output)
